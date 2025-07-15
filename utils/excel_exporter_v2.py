#!/usr/bin/env python3
# -*- coding: utf-8 -*-

try:
    from PySide6.QtCore import QObject
    PYSIDE6_AVAILABLE = True
except ImportError:
    PYSIDE6_AVAILABLE = False
    # 提供备用实现
    class QObject:
        def __init__(self, *args, **kwargs):
            pass

"""
TimeNest Excel导出工具 v2
支持多种格式和样式的课程表导出
"""

import logging
import pandas as pd
from typing import Dict, List, Any, Optional, Union
from pathlib import Path
from datetime import datetime, time
from PySide6.QtCore import QObject, Signal, QThread
from dataclasses import dataclass
from enum import Enum

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from models.schedule import Schedule


class ExportFormat(Enum):
    """导出格式"""
    SIMPLE = "simple"
    DETAILED = "detailed"
    WEEKLY = "weekly"
    DAILY = "daily"
    TIMETABLE = "timetable"


class ExportStyle(Enum):
    """导出样式"""
    BASIC = "basic"
    COLORFUL = "colorful"
    PROFESSIONAL = "professional"
    MINIMAL = "minimal"


@dataclass
class ExportOptions:
    """导出选项"""
    format_type: ExportFormat = ExportFormat.WEEKLY
    style: ExportStyle = ExportStyle.COLORFUL
    include_headers: bool = True
    include_time: bool = True
    include_teacher: bool = True
    include_classroom: bool = True
    include_description: bool = False
    merge_cells: bool = True
    auto_width: bool = True
    freeze_panes: bool = True
    add_borders: bool = True
    highlight_current_time: bool = False
    custom_colors: Dict[str, str] = None
    # 新增打印优化选项
    print_optimized: bool = True
    page_orientation: str = "landscape"  # landscape 或 portrait
    add_statistics: bool = False
    split_by_week: bool = False
    add_header_info: bool = True
    include_week_range: bool = True

    def __post_init__(self):
        if self.custom_colors is None:
            self.custom_colors = {}


class ExcelExportThread(QThread):
    """Excel导出线程"""
    progress_updated = Signal(int)
    export_finished = Signal(bool, str)
    
    def __init__(self, schedule: Schedule, file_path: str, options: ExportOptions = None):
        super().__init__()
        self.logger = logging.getLogger(f'{__name__}.ExcelExportThread')
        self.schedule = schedule
        self.file_path = file_path
        self.options = options or ExportOptions()
    
    def run(self):
        """执行导出"""
        try:
            exporter = ExcelExporterV2()
            exporter.progress_updated.connect(self.progress_updated.emit)
            
            success = exporter.export_schedule(
                schedule=self.schedule,
                file_path=self.file_path,
                options=self.options
            )
            
            
            if success:
                self.export_finished.emit(True, f"导出成功: {self.file_path}")
            else:
                self.export_finished.emit(False, "导出失败")
                
        except Exception as e:
            error_msg = f"导出线程出错: {e}"
            self.logger.error(error_msg, exc_info=True)
            self.export_finished.emit(False, error_msg)


class ExcelExporterV2(QObject):
    """Excel导出工具类 v2"""
    progress_updated = Signal(int)
    
    def __init__(self):
        super().__init__()
        self.logger = logging.getLogger(__name__)
        
        # 预定义颜色方案
        self.color_schemes = {
            'default': {
                'header': 'FF4472C4',
                'subject_colors': [
                    'FFE7E6FF', 'FFFCE4EC', 'FFE8F5E8', 'FFFFF3E0',
                    'FFE3F2FD', 'FFF3E5F5', 'FFFFE0B2', 'FFEFEBE9'
                ]
            },
            'professional': {
                'header': 'FF2F4F4F',
                'subject_colors': [
                    'FFF5F5F5', 'FFE8E8E8', 'FFDCDCDC', 'FFD3D3D3'
                ]
            },
            'colorful': {
                'header': 'FF1976D2',
                'subject_colors': [
                    'FFBBDEFB', 'FFC8E6C9', 'FFFFCDD2', 'FFFFD54F',
                    'FFCE93D8', 'FF80CBC4', 'FFFFAB91', 'FFA5D6A7'
                ]
            }
        }
        
        # 周几映射
        self.weekday_names = {
            'monday': '周一',
            'tuesday': '周二', 
            'wednesday': '周三',
            'thursday': '周四',
            'friday': '周五',
            'saturday': '周六',
            'sunday': '周日'
        }
    
    def export_schedule(self, schedule: Schedule, file_path: str, options: ExportOptions = None) -> bool:
        """导出课程表到Excel文件"""
        try:
            if not OPENPYXL_AVAILABLE:
                self.logger.error("openpyxl库未安装，无法导出Excel文件")
                return False
            
            
            if options is None:
                options = ExportOptions()
            
            self.progress_updated.emit(10)
            
            # 根据格式类型选择导出方法
            if options.format_type == ExportFormat.WEEKLY:
                success = self._export_weekly_format(schedule, file_path, options)
            elif options.format_type == ExportFormat.TIMETABLE:
                success = self._export_timetable_format(schedule, file_path, options)
            elif options.format_type == ExportFormat.DETAILED:
                success = self._export_detailed_format(schedule, file_path, options)
            else:  # SIMPLE
                success = self._export_simple_format(schedule, file_path, options)
            
            self.progress_updated.emit(100)
            
            
            if success:
                self.logger.info(f"课程表导出成功: {file_path}")
            else:
                self.logger.error("课程表导出失败")
            
            return success
            
        except Exception as e:
            self.logger.error(f"导出课程表失败: {e}", exc_info=True)
            return False
    
    def _export_weekly_format(self, schedule: Schedule, file_path: str, options: ExportOptions) -> bool:
        """导出周视图格式"""
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "课程表"
            
            self.progress_updated.emit(20)
            
            # 创建周视图数据结构
            weekdays = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
            time_slots = sorted(schedule.time_slots, key=lambda x: x.start_time)
            
            # 构建表头
            headers = ['时间']
            headers.extend([self.weekday_names[day] for day in weekdays])
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                if options.style != ExportStyle.MINIMAL:
                    self._apply_header_style(cell, options)
            
            self.progress_updated.emit(40)
            
            # 构建课程数据
            row = 2
            for time_slot in time_slots:
                # 时间列
                time_str = f"{time_slot.start_time.strftime('%H:%M')}-{time_slot.end_time.strftime('%H:%M')}"
                worksheet.cell(row=row, column=1, value=time_str)
                
                # 各天的课程
                for col, weekday in enumerate(weekdays, 2):
                    class_items = [c for c in schedule.classes 
                                 if c.weekday == weekday and c.time_slot_id == time_slot.id and c.is_active]
                    
                    if class_items:
                        class_item = class_items[0]
                        subject = schedule.get_subject(class_item.subject_id)
                        
                        # 构建单元格内容
                        cell_content = []
                        if subject:
                            cell_content.append(subject.name)
                            
                            
                        if options.include_teacher and class_item.teacher:
                            cell_content.append(f"教师: {class_item.teacher}")
                        elif options.include_teacher and subject and subject.teacher:
                            cell_content.append(f"教师: {subject.teacher}")
                            
                            
                        if options.include_classroom and class_item.classroom:
                            cell_content.append(f"教室: {class_item.classroom}")
                        
                        cell_value = '\n'.join(cell_content)
                        cell = worksheet.cell(row=row, column=col, value=cell_value)
                        
                        # 应用样式
                        if options.style != ExportStyle.MINIMAL and subject:
                            self._apply_subject_style(cell, subject, options)
                    else:
                        worksheet.cell(row=row, column=col, value="")
                
                row += 1
            
            self.progress_updated.emit(70)
            
            # 应用格式设置
            self._apply_worksheet_formatting(worksheet, options)
            
            self.progress_updated.emit(90)
            
            # 保存文件
            workbook.save(file_path)
            return True
            
        except Exception as e:
            self.logger.error(f"导出周视图格式失败: {e}")
            return False
    
    def _export_timetable_format(self, schedule: Schedule, file_path: str, options: ExportOptions) -> bool:
        """导出时间表格式"""
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = f"{schedule.name} - 课程表"
            
            # 添加标题
            title_cell = worksheet.cell(row=1, column=1, value=f"{schedule.name}")
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            worksheet.merge_cells('A1:H1')
            
            self.progress_updated.emit(30)
            
            # 创建课程表网格
            weekdays = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday']
            time_slots = sorted(schedule.time_slots, key=lambda x: x.start_time)
            
            # 表头行
            headers = ['节次/时间'] + [self.weekday_names[day] for day in weekdays]
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=3, column=col, value=header)
                self._apply_header_style(cell, options)
            
            # 数据行
            for row_idx, time_slot in enumerate(time_slots, 4):
                # 节次和时间
                time_info = f"第{row_idx-3}节\n{time_slot.start_time.strftime('%H:%M')}-{time_slot.end_time.strftime('%H:%M')}"
                time_cell = worksheet.cell(row=row_idx, column=1, value=time_info)
                time_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # 各天课程
                for col_idx, weekday in enumerate(weekdays, 2):
                    class_items = [c for c in schedule.classes 
                                 if c.weekday == weekday and c.time_slot_id == time_slot.id and c.is_active]
                    
                    if class_items:
                        class_item = class_items[0]
                        subject = schedule.get_subject(class_item.subject_id)
                        
                        
                        if subject:
                            content_parts = [subject.name]
                            
                            
                            if options.include_teacher:
                                teacher = class_item.teacher or subject.teacher
                                if teacher:
                                    content_parts.append(teacher)
                            
                            
                            if options.include_classroom and class_item.classroom:
                                content_parts.append(class_item.classroom)
                            
                            cell_content = '\n'.join(content_parts)
                            cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_content)
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            
                            # 应用科目颜色
                            if options.style == ExportStyle.COLORFUL:
                                self._apply_subject_style(cell, subject, options)
                    else:
                        cell = worksheet.cell(row=row_idx, column=col_idx, value="")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            self.progress_updated.emit(80)
            
            # 设置行高和列宽
            for row in worksheet.iter_rows(min_row=4, max_row=3+len(time_slots)):
                worksheet.row_dimensions[row[0].row].height = 60
            
            worksheet.column_dimensions.get('A').width = 15
            for col in range(2, 7):
                worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = 20
            
            # 添加边框
            if options.add_borders:
                self._add_borders(worksheet, 3, 1, 3+len(time_slots), 6)
            
            # 保存文件
            workbook.save(file_path)
            return True
            
        except Exception as e:
            self.logger.error(f"导出时间表格式失败: {e}")
            return False

    def _export_simple_format(self, schedule: Schedule, file_path: str, options: ExportOptions) -> bool:
        """导出简单格式"""
        try:
            # 使用pandas创建简单的DataFrame
            data = []
            for class_item in schedule.classes:
                if not class_item.is_active:
                    continue

                subject = schedule.get_subject(class_item.subject_id)
                time_slot = schedule.get_time_slot(class_item.time_slot_id)

                row = {
                    '星期': self.weekday_names.get(class_item.weekday, class_item.weekday),
                    '时间': f"{time_slot.start_time.strftime('%H:%M')}-{time_slot.end_time.strftime('%H:%M')}" if time_slot else "",
                    '科目': subject.name if subject else "",
                    '教师': class_item.teacher or (subject.teacher if subject else ""),
                    '教室': class_item.classroom or ""
                }
                data.append(row)

            df = pd.DataFrame(data)
            df.to_excel(file_path, index=False, sheet_name=schedule.name)
            return True

        except Exception as e:
            self.logger.error(f"导出简单格式失败: {e}")
            return False

    def _export_detailed_format(self, schedule: Schedule, file_path: str, options: ExportOptions) -> bool:
        """导出详细格式"""
        try:
            workbook = Workbook()

            # 课程表概览
            overview_sheet = workbook.active
            overview_sheet.title = "课程表概览"

            # 添加课程表信息
            overview_sheet.cell(row=1, column=1, value="课程表名称")
            overview_sheet.cell(row=1, column=2, value=schedule.name)
            overview_sheet.cell(row=2, column=1, value="创建时间")
            overview_sheet.cell(row=2, column=2, value=schedule.created_date.strftime('%Y-%m-%d %H:%M:%S'))
            overview_sheet.cell(row=3, column=1, value="科目数量")
            overview_sheet.cell(row=3, column=2, value=len(schedule.subjects))
            overview_sheet.cell(row=4, column=1, value="课程数量")
            overview_sheet.cell(row=4, column=2, value=len([c for c in schedule.classes if c.is_active]))

            # 科目列表
            subjects_sheet = workbook.create_sheet("科目列表")
            subjects_headers = ['科目名称', '教师', '颜色', '描述']
            for col, header in enumerate(subjects_headers, 1):
                subjects_sheet.cell(row=1, column=col, value=header)

            for row, subject in enumerate(schedule.subjects, 2):
                subjects_sheet.cell(row=row, column=1, value=subject.name)
                subjects_sheet.cell(row=row, column=2, value=subject.teacher)
                subjects_sheet.cell(row=row, column=3, value=subject.color)
                subjects_sheet.cell(row=row, column=4, value=subject.description)

            # 课程详情
            classes_sheet = workbook.create_sheet("课程详情")
            classes_headers = ['星期', '时间段', '科目', '教师', '教室', '备注']
            for col, header in enumerate(classes_headers, 1):
                classes_sheet.cell(row=1, column=col, value=header)

            row = 2
            for class_item in schedule.classes:
                if not class_item.is_active:
                    continue

                subject = schedule.get_subject(class_item.subject_id)
                time_slot = schedule.get_time_slot(class_item.time_slot_id)

                classes_sheet.cell(row=row, column=1, value=self.weekday_names.get(class_item.weekday, class_item.weekday))
                classes_sheet.cell(row=row, column=2, value=time_slot.name if time_slot else "")
                classes_sheet.cell(row=row, column=3, value=subject.name if subject else "")
                classes_sheet.cell(row=row, column=4, value=class_item.teacher or (subject.teacher if subject else ""))
                classes_sheet.cell(row=row, column=5, value=class_item.classroom)
                classes_sheet.cell(row=row, column=6, value=class_item.description)
                row += 1

            # 保存文件
            workbook.save(file_path)
            return True

        except Exception as e:
            self.logger.error(f"导出详细格式失败: {e}")
            return False

    def _apply_header_style(self, cell, options: ExportOptions):
        """应用表头样式"""
        try:
            if options.style == ExportStyle.MINIMAL:
                return

            # 字体
            cell.font = Font(bold=True, color='FFFFFF')

            # 背景色
            scheme = self.color_schemes.get('colorful' if options.style == ExportStyle.COLORFUL else 'professional')
            header_color = scheme.get('header')
            cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')

            # 对齐
            cell.alignment = Alignment(horizontal='center', vertical='center')

        except Exception as e:
            self.logger.error(f"应用表头样式失败: {e}")

    def _apply_subject_style(self, cell, subject, options: ExportOptions):
        """应用科目样式"""
        try:
            if options.style == ExportStyle.MINIMAL:
                return

            # 获取颜色方案
            scheme = self.color_schemes.get('colorful' if options.style == ExportStyle.COLORFUL else 'professional')
            colors = scheme.get('subject_colors')

            # 根据科目名称选择颜色
            color_index = hash(subject.name) % len(colors)
            bg_color = colors[color_index]

            # 应用背景色
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

            # 对齐和换行
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        except Exception as e:
            self.logger.error(f"应用科目样式失败: {e}")

    def _apply_worksheet_formatting(self, worksheet, options: ExportOptions):
        """应用工作表格式"""
        try:
            # 自动调整列宽
            if options.auto_width:
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 30)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            # 冻结窗格
            if options.freeze_panes:
                worksheet.freeze_panes = 'B2'

            # 添加边框
            if options.add_borders:
                self._add_borders(worksheet, 1, 1, worksheet.max_row, worksheet.max_column)

        except Exception as e:
            self.logger.error(f"应用工作表格式失败: {e}")

    def _add_borders(self, worksheet, start_row: int, start_col: int, end_row: int, end_col: int):
        """添加边框"""
        try:
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    worksheet.cell(row=row, column=col).border = thin_border

        except Exception as e:
            self.logger.error(f"添加边框失败: {e}")

    def export_schedule_async(self, schedule: Schedule, file_path: str, options: ExportOptions = None) -> ExcelExportThread:
        """异步导出课程表"""
        export_thread = ExcelExportThread(schedule, file_path, options)
        return export_thread
