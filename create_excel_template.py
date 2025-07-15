#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
创建Excel课程表模板
"""

import pandas as pd
from pathlib import Path

def create_excel_template():
    """创建Excel课程表模板"""
    
    # 时间段
    time_slots = [
        "08:00-08:45", "08:55-09:40", "10:00-10:45", "10:55-11:40",
        "14:00-14:45", "14:55-15:40", "16:00-16:45", "16:55-17:40", 
        "19:00-19:45", "19:55-20:40", "20:50-21:35"
    ]
    
    # 星期
    weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    
    # 创建数据结构
    data = {"时间": time_slots}
    for day in weekdays:
        data[day] = [""] * len(time_slots)
    
    # 添加示例课程数据
    # 格式：课程名称;教师姓名;上课地点;周次信息
    data["周一"][0] = "高等数学;张教授;教学楼A101;1-16周"
    data["周一"][1] = "线性代数;李教授;教学楼B203;1-8周;概率论;李教授;教学楼B203;10-16周"
    data["周二"][2] = "大学英语;王教授;外语楼301;1-16周"
    data["周三"][0] = "计算机程序设计;赵教授;实验楼501;2-16周"
    data["周三"][3] = "数据结构;赵教授;实验楼501;1-8周;算法设计;赵教授;实验楼501;9-16周"
    data["周四"][3] = "大学物理;陈教授;理科楼201;1-12周"
    data["周五"][1] = "体育;刘教练;体育馆;1-16周"
    data["周五"][4] = "思想政治;孙教授;人文楼101;1-16周(单)"
    data["周六"][0] = "选修课;各教授;综合楼;3,5,7,9,11,13,15周"
    
    # 创建DataFrame
    df = pd.DataFrame(data)
    
    # 保存为Excel文件
    template_path = "schedule_template.xlsx"
    
    try:
        with pd.ExcelWriter(template_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='课程表', index=False)
            
            # 获取工作表进行格式化
            worksheet = writer.sheets['课程表']
            
            # 设置列宽
            worksheet.column_dimensions['A'].width = 15  # 时间列
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                worksheet.column_dimensions[col].width = 30  # 星期列
            
            # 设置行高
            for row in range(1, len(time_slots) + 2):
                worksheet.row_dimensions[row].height = 40
            
            # 添加边框和对齐
            from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'), 
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            center_alignment = Alignment(
                horizontal='center', 
                vertical='center', 
                wrap_text=True
            )
            
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            
            # 格式化所有单元格
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = center_alignment
                    
                    # 设置表头样式
                    if cell.row == 1:
                        cell.font = header_font
                        cell.fill = header_fill
        
        print(f"✅ Excel模板已创建: {template_path}")
        return True
        
    except Exception as e:
        print(f"❌ 创建Excel模板失败: {e}")
        return False

if __name__ == "__main__":
    create_excel_template()
