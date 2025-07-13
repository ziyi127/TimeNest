#!/usr/bin/env python3
"""
增强的Excel导出模块

提供课程表和其他数据的Excel导出功能
"""

import logging
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlsxwriter
    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False


class ExcelExportError(Exception):
    """Excel导出异常"""
    pass


class ScheduleExporter:
    """课程表导出器"""
    
    def __init__(self):
        self.logger = logging.getLogger(f'{__name__}.ScheduleExporter')
        
        if not (OPENPYXL_AVAILABLE or XLSXWRITER_AVAILABLE):
            self.logger.warning("Excel导出库不可用，将使用基础CSV导出")
    
    def export_schedule_to_excel(self, schedule_data: List[Dict[str, Any]], 
                                export_path: str, 
                                format_type: str = "detailed") -> bool:
        """
        导出课程表到Excel文件
        
        Args:
            schedule_data: 课程数据列表
            export_path: 导出文件路径
            format_type: 导出格式类型 ("detailed", "simple", "weekly")
        
        Returns:
            bool: 导出是否成功
        """
        try:
            if not schedule_data:
                raise ExcelExportError("没有课程数据可导出")
            
            export_path = Path(export_path)
            
            if OPENPYXL_AVAILABLE:
                return self._export_with_openpyxl(schedule_data, export_path, format_type)
            elif XLSXWRITER_AVAILABLE:
                return self._export_with_xlsxwriter(schedule_data, export_path, format_type)
            else:
                return self._export_to_csv(schedule_data, export_path)
                
        except Exception as e:
            self.logger.error(f"Excel导出失败: {e}")
            raise ExcelExportError(f"导出失败: {e}")
    
    def _export_with_openpyxl(self, schedule_data: List[Dict[str, Any]], 
                             export_path: Path, format_type: str) -> bool:
        """使用openpyxl导出"""
        try:
            # 创建工作簿
            workbook = openpyxl.Workbook()
            
            if format_type == "weekly":
                self._create_weekly_schedule(workbook, schedule_data)
            elif format_type == "detailed":
                self._create_detailed_schedule(workbook, schedule_data)
            else:
                self._create_simple_schedule(workbook, schedule_data)
            
            # 保存文件
            workbook.save(export_path)
            self.logger.info(f"Excel文件已保存: {export_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"openpyxl导出失败: {e}")
            return False
    
    def _create_weekly_schedule(self, workbook, schedule_data):
        """创建周课程表格式"""
        try:
            ws = workbook.active
            ws.title = "周课程表"
            
            # 设置表头
            days = ['时间', '周一', '周二', '周三', '周四', '周五', '周六', '周日']
            for col, day in enumerate(days, 1):
                cell = ws.cell(row=1, column=col, value=day)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 生成时间段
            time_slots = self._generate_time_slots()
            
            # 创建课程表网格
            schedule_grid = {}
            for course in schedule_data:
                day = course.get('day_of_week', 0)
                start_time = course.get('start_time', '')
                course_name = course.get('name', '未知课程')
                location = course.get('location', '')
                
                if day in range(1, 8):  # 周一到周日
                    key = f"{day}_{start_time}"
                    schedule_grid[key] = f"{course_name}\n{location}"
            
            # 填充数据
            for row, time_slot in enumerate(time_slots, 2):
                ws.cell(row=row, column=1, value=time_slot)
                
                for day in range(1, 8):
                    key = f"{day}_{time_slot}"
                    course_info = schedule_grid.get(key, '')
                    ws.cell(row=row, column=day + 1, value=course_info)
            
            # 设置列宽
            for col in range(1, 9):
                ws.column_dimensions[chr(64 + col)].width = 15
                
        except Exception as e:
            self.logger.error(f"创建周课程表失败: {e}")
    
    def _create_detailed_schedule(self, workbook, schedule_data):
        """创建详细课程表格式"""
        try:
            ws = workbook.active
            ws.title = "详细课程表"
            
            # 设置表头
            headers = ['课程名称', '教师', '地点', '星期', '开始时间', '结束时间', '学分', '备注']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 填充数据
            for row, course in enumerate(schedule_data, 2):
                ws.cell(row=row, column=1, value=course.get('name', ''))
                ws.cell(row=row, column=2, value=course.get('teacher', ''))
                ws.cell(row=row, column=3, value=course.get('location', ''))
                ws.cell(row=row, column=4, value=self._get_day_name(course.get('day_of_week', 0)))
                ws.cell(row=row, column=5, value=course.get('start_time', ''))
                ws.cell(row=row, column=6, value=course.get('end_time', ''))
                ws.cell(row=row, column=7, value=course.get('credits', ''))
                ws.cell(row=row, column=8, value=course.get('notes', ''))
            
            # 设置列宽
            column_widths = [20, 15, 15, 10, 12, 12, 8, 25]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col)].width = width
                
        except Exception as e:
            self.logger.error(f"创建详细课程表失败: {e}")
    
    def _create_simple_schedule(self, workbook, schedule_data):
        """创建简单课程表格式"""
        try:
            ws = workbook.active
            ws.title = "课程表"
            
            # 转换为DataFrame
            df = pd.DataFrame(schedule_data)
            
            # 选择主要列
            main_columns = ['name', 'teacher', 'location', 'day_of_week', 'start_time', 'end_time']
            available_columns = [col for col in main_columns if col in df.columns]
            
            if available_columns:
                df_export = df[available_columns]
                
                # 重命名列
                column_names = {
                    'name': '课程名称',
                    'teacher': '教师',
                    'location': '地点',
                    'day_of_week': '星期',
                    'start_time': '开始时间',
                    'end_time': '结束时间'
                }
                df_export = df_export.rename(columns=column_names)
                
                # 转换星期数字为中文
                if '星期' in df_export.columns:
                    df_export['星期'] = df_export['星期'].apply(self._get_day_name)
                
                # 写入Excel
                for r in dataframe_to_rows(df_export, index=False, header=True):
                    ws.append(r)
                
                # 设置表头样式
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
        except Exception as e:
            self.logger.error(f"创建简单课程表失败: {e}")
    
    def _export_with_xlsxwriter(self, schedule_data: List[Dict[str, Any]], 
                               export_path: Path, format_type: str) -> bool:
        """使用xlsxwriter导出"""
        try:
            workbook = xlsxwriter.Workbook(str(export_path))
            worksheet = workbook.add_worksheet('课程表')
            
            # 设置格式
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#CCCCCC',
                'border': 1
            })
            
            # 写入表头
            headers = ['课程名称', '教师', '地点', '星期', '开始时间', '结束时间']
            for col, header in enumerate(headers):
                worksheet.write(0, col, header, header_format)
            
            # 写入数据
            for row, course in enumerate(schedule_data, 1):
                worksheet.write(row, 0, course.get('name', ''))
                worksheet.write(row, 1, course.get('teacher', ''))
                worksheet.write(row, 2, course.get('location', ''))
                worksheet.write(row, 3, self._get_day_name(course.get('day_of_week', 0)))
                worksheet.write(row, 4, course.get('start_time', ''))
                worksheet.write(row, 5, course.get('end_time', ''))
            
            workbook.close()
            self.logger.info(f"Excel文件已保存: {export_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"xlsxwriter导出失败: {e}")
            return False
    
    def _export_to_csv(self, schedule_data: List[Dict[str, Any]], export_path: Path) -> bool:
        """导出为CSV文件（备用方案）"""
        try:
            df = pd.DataFrame(schedule_data)
            
            # 转换星期数字为中文
            if 'day_of_week' in df.columns:
                df['day_of_week'] = df['day_of_week'].apply(self._get_day_name)
            
            # 保存为CSV
            csv_path = export_path.with_suffix('.csv')
            df.to_csv(csv_path, index=False, encoding='utf-8-sig')
            
            self.logger.info(f"CSV文件已保存: {csv_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"CSV导出失败: {e}")
            return False
    
    def _generate_time_slots(self) -> List[str]:
        """生成时间段列表"""
        time_slots = []
        start_hour = 8
        end_hour = 22
        
        for hour in range(start_hour, end_hour):
            time_slots.append(f"{hour:02d}:00")
            time_slots.append(f"{hour:02d}:30")
        
        return time_slots
    
    def _get_day_name(self, day_num: int) -> str:
        """获取星期名称"""
        day_names = {
            0: '周日',
            1: '周一', 
            2: '周二',
            3: '周三',
            4: '周四',
            5: '周五',
            6: '周六',
            7: '周日'
        }
        return day_names.get(day_num, '未知')


# 导出函数，供其他模块使用
def export_schedule_to_excel(schedule_data: List[Dict[str, Any]], 
                           export_path: str, 
                           format_type: str = "detailed") -> bool:
    """
    导出课程表到Excel文件
    
    Args:
        schedule_data: 课程数据列表
        export_path: 导出文件路径
        format_type: 导出格式类型
    
    Returns:
        bool: 导出是否成功
    """
    exporter = ScheduleExporter()
    return exporter.export_schedule_to_excel(schedule_data, export_path, format_type)
